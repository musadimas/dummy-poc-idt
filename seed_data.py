#!/usr/bin/env python3
"""
seed_data.py — EDC POC database seeder
Reads poi_edc.csv and populates all tables with realistic Indonesian payment data.

Requirements: psycopg2-binary
    pip install psycopg2-binary

Override DB connection via environment variables:
    PG_HOST, PG_PORT, PG_DB, PG_USER, PG_PASSWORD
"""

import csv
import math
import os
import random
import re
import string
import sys
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from pathlib import Path

import psycopg2
from psycopg2.extras import execute_values

from config import *
from helpers import *

# ──────────────────────────────────────────────────────────────────────────────
# DATABASE HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def reset_database(conn) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            TRUNCATE
                transaction_log,
                settlement,
                transactions,
                terminals,
                cards,
                merchants,
                admin_areas
            RESTART IDENTITY CASCADE
        """)
    conn.commit()
    print("      All tables truncated.")


# ──────────────────────────────────────────────────────────────────────────────
# SCHEMA MIGRATION HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _ensure_merchant_columns(conn) -> None:
    """Add reference and merchant_status columns to merchants if not already present."""
    with conn.cursor() as cur:
        cur.execute("""
            ALTER TABLE merchants
                ADD COLUMN IF NOT EXISTS reference       VARCHAR(255),
                ADD COLUMN IF NOT EXISTS merchant_status VARCHAR(20) DEFAULT 'ACTIVE'
        """)
    conn.commit()


def _sync_merchant_data(conn) -> None:
    """
    For all 'delete' rows in list_edc.csv:
      1. Normalize closed_from (quarter format or empty → random ISO date).
      2. Write updated values back to the CSV.
      3. In DB: set merchant_status = 'INACTIVE' for matched merchants if not already.
    Also updates the reference column for all merchants that have a CSV id.
    """
    import csv as _csv_mod

    csv_path = _BATCH_UNIFIED_CSV
    if not csv_path.exists():
        return

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader    = _csv_mod.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows      = list(reader)

    csv_changed = 0
    for row in rows:
        if row.get("report_status", "").lower() != "delete":
            continue
        original = row.get("closed_from", "").strip()
        parsed   = _parse_closed_from(original)
        if not parsed:
            # No date at all — generate a random date in the past year
            _end   = date.today()
            _start = _end - timedelta(days=365)
            parsed = (_start + timedelta(days=random.randint(0, (_end - _start).days))).isoformat()
        if parsed != original:
            row["closed_from"] = parsed
            csv_changed += 1

    if csv_changed:
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = _csv_mod.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        print(f"      [sync] Normalised {csv_changed} closed_from date(s) in {csv_path.name}")

    # Build name→(id, closed_from) lookup for all rows that have an id
    ref_map: dict[str, str] = {}
    inactive_names: set[str] = set()
    for row in rows:
        name = row.get("name", "").strip()
        if not name:
            continue
        csv_id = row.get("id", "").strip()
        if csv_id:
            ref_map[name] = csv_id
        if row.get("report_status", "").lower() == "delete" and row.get("closed_from", "").strip():
            inactive_names.add(name)

    if not ref_map and not inactive_names:
        return

    with conn.cursor() as cur:
        cur.execute("SELECT merchant_name, merchant_id FROM merchants")
        db_map = {n: mid for n, mid in cur.fetchall()}

    updated_ref = updated_status = 0
    with conn.cursor() as cur:
        for name, csv_id in ref_map.items():
            m_id = db_map.get(name)
            if m_id is None:
                continue
            cur.execute(
                "UPDATE merchants SET reference = %s WHERE merchant_id = %s AND (reference IS NULL OR reference != %s)",
                (csv_id, m_id, csv_id),
            )
            updated_ref += cur.rowcount

        for name in inactive_names:
            m_id = db_map.get(name)
            if m_id is None:
                continue
            cur.execute(
                "UPDATE merchants SET merchant_status = 'INACTIVE' "
                "WHERE merchant_id = %s AND (merchant_status IS NULL OR merchant_status != 'INACTIVE')",
                (m_id,),
            )
            updated_status += cur.rowcount

    conn.commit()
    if updated_ref:
        print(f"      [sync] Updated reference for {updated_ref} merchant(s)")
    if updated_status:
        print(f"      [sync] Marked {updated_status} merchant(s) as INACTIVE")


# ──────────────────────────────────────────────────────────────────────────────
# LOAD STATIC REFERENCE DATA
# ──────────────────────────────────────────────────────────────────────────────

def load_acquirers(conn) -> dict[str, int]:
    """Return {bank_code: acquirer_id}. Static data seeded by schema.sql."""
    with conn.cursor() as cur:
        cur.execute("SELECT bank_code, acquirer_id FROM acquirers")
        result = {r[0]: r[1] for r in cur.fetchall()}
    if not result:
        raise RuntimeError("No acquirers found — run schema.sql first.")
    return result


def load_qris_issuers(conn) -> dict[str, int]:
    """Return {issuer_code: qris_issuer_id}. Static data seeded by schema.sql."""
    with conn.cursor() as cur:
        cur.execute("SELECT issuer_code, qris_issuer_id FROM qris_issuers")
        result = {r[0]: r[1] for r in cur.fetchall()}
    if not result:
        raise RuntimeError("No QRIS issuers found — run schema.sql first.")
    return result


# ──────────────────────────────────────────────────────────────────────────────
# INSERT ADMIN AREAS  (Province → City/Regency → District → Village)
# ──────────────────────────────────────────────────────────────────────────────

def insert_admin_areas(conn, csv_rows: list[dict]) -> dict[tuple, int]:
    """
    Build the Province → City/Regency → District hierarchy from CSV admin columns.
    Returns a mapping of (province, city, district) → area_id for the deepest
    available level (used as the merchant FK).

    Structure:
      Level 1 = Province  (Provinsi)
      Level 2 = City/Regency  (Kota/Kabupaten)
      Level 3 = District  (Kecamatan)
      Level 4 = Village  (Kelurahan/Desa) — structure ready, not in CSV data
    """
    # Collect unique tuples present in the CSV
    combos: set[tuple[str | None, str | None, str | None]] = set()
    for row in csv_rows:
        combos.add(_resolve_admin_tuple(row))

    province_ids: dict[str, int] = {}   # name → area_id
    city_ids:     dict[tuple, int] = {} # (province, city) → area_id
    district_ids: dict[tuple, int] = {} # (province, city, district) → area_id

    with conn.cursor() as cur:
        # ── Level 1: provinces ──────────────────────────────────────────────
        provinces = {p for p, c, d in combos if p}
        for prov in sorted(provinces):
            cur.execute(
                """
                INSERT INTO admin_areas (area_name, area_level, parent_id)
                VALUES (%s, 1, NULL)
                ON CONFLICT (area_name, area_level, parent_id) DO UPDATE
                    SET area_name = EXCLUDED.area_name
                RETURNING area_id
                """,
                (prov,),
            )
            province_ids[prov] = cur.fetchone()[0]

        # ── Level 2: cities ─────────────────────────────────────────────────
        cities = {(p, c) for p, c, d in combos if p and c}
        for prov, city in sorted(cities):
            parent = province_ids.get(prov)
            cur.execute(
                """
                INSERT INTO admin_areas (area_name, area_level, parent_id)
                VALUES (%s, 2, %s)
                ON CONFLICT (area_name, area_level, parent_id) DO UPDATE
                    SET area_name = EXCLUDED.area_name
                RETURNING area_id
                """,
                (city, parent),
            )
            city_ids[(prov, city)] = cur.fetchone()[0]

        # ── Level 3: districts ───────────────────────────────────────────────
        for prov, city, dist in sorted(combos):
            if not dist:
                continue
            parent = city_ids.get((prov, city)) or province_ids.get(prov)
            cur.execute(
                """
                INSERT INTO admin_areas (area_name, area_level, parent_id)
                VALUES (%s, 3, %s)
                ON CONFLICT (area_name, area_level, parent_id) DO UPDATE
                    SET area_name = EXCLUDED.area_name
                RETURNING area_id
                """,
                (dist, parent),
            )
            district_ids[(prov, city, dist)] = cur.fetchone()[0]

    conn.commit()

    # Build merchant lookup: (prov, city, dist) → deepest area_id
    result: dict[tuple, int] = {}
    for prov, city, dist in combos:
        if dist and (prov, city, dist) in district_ids:
            result[(prov, city, dist)] = district_ids[(prov, city, dist)]
        elif city and (prov, city) in city_ids:
            result[(prov, city, dist)] = city_ids[(prov, city)]
        elif prov and prov in province_ids:
            result[(prov, city, dist)] = province_ids[prov]

    return result


# ──────────────────────────────────────────────────────────────────────────────
# INSERT MERCHANTS + TERMINALS
# ──────────────────────────────────────────────────────────────────────────────

def insert_merchants_and_terminals(
    conn,
    csv_rows: list[dict],
    acquirer_ids: dict[str, int],
    area_map: dict[tuple, int],
    idx_offset: int = 0,
) -> list[dict]:
    """
    Insert one merchant + one terminal per active CSV row.
    Returns list of info dicts consumed by the transaction generator.
    """
    acquirer_codes = list(acquirer_ids.keys())
    merchant_insert_rows = []
    terminal_meta        = []

    for idx, row in enumerate(csv_rows):
        category      = row.get("primarycategorynm", "").strip()
        mcc           = CATEGORY_TO_MCC.get(category, "5999")
        # Registration date: random within 3–18 months before DATE_START
        reg_date      = (DATE_START - timedelta(days=random.randint(90, 548))).strftime("%Y%m%d")
        serial_number = f"SN-2026-{idx_offset + idx + 1:04d}"
        merchant_code = f"MCH-{mcc}-{reg_date}-{idx_offset + idx + 1:05d}"
        terminal_code = f"TID-{mcc}-{reg_date}-{idx_offset + idx + 1:05d}"

        lat_str = row.get("displaylatitude",  "").strip()
        lon_str = row.get("displaylongitude", "").strip()
        latitude  = float(lat_str)  if lat_str  else None
        longitude = float(lon_str) if lon_str else None

        acquirer_id   = acquirer_ids[random.choice(acquirer_codes)]
        admin_tuple   = _resolve_admin_tuple(row)
        admin_area_id = area_map.get(admin_tuple)

        address = " ".join(filter(None, [
            row.get("hno", "").strip(),
            row.get("streetname", "").strip(),
        ])) or None
        # Denormalised display city: prefer City level, fall back to Province
        city = admin_tuple[1] or admin_tuple[0] or "Jakarta"

        reference = row.get("_xlsx_id", "") or None
        merchant_insert_rows.append((
            row.get("name1", f"Merchant {idx+1}").strip(),
            merchant_code,
            mcc,
            address,
            city,
            admin_area_id,
            latitude,
            longitude,
            acquirer_id,
            reference,
        ))

        sched_override = row.get("_schedule_override")
        schedule = sched_override if sched_override else {wd: resolve_hours(row, wd, category) for wd in range(7)}

        cf_str     = row.get("closed_from", "")
        if isinstance(cf_str, str):
            cf_str = cf_str.strip()
        else:
            cf_str = ""
        closed_from = date.fromisoformat(cf_str) if cf_str else None

        terminal_meta.append({
            "merchant_name": row.get("name1", f"Merchant {idx+1}").strip(),
            "merchant_code": merchant_code,
            "terminal_code": terminal_code,
            "serial_number": serial_number,
            "model":         random.choice(TERMINAL_MODELS),
            "category":      category,
            "schedule":      schedule,
            "closed_from":   closed_from,
        })

    # Insert merchants, capture generated UUIDs (fetch=True collects all pages)
    with conn.cursor() as cur:
        returned = execute_values(
            cur,
            """
            INSERT INTO merchants
                (merchant_name, merchant_code, mcc_code, address, city,
                 admin_area_id, latitude, longitude, acquirer_id, reference)
            VALUES %s
            RETURNING merchant_id, merchant_code
            """,
            merchant_insert_rows,
            page_size=200,
            fetch=True,
        )
        merchant_code_to_uuid = {r[1]: r[0] for r in returned}

    # Insert terminals, capture generated UUIDs (fetch=True collects all pages)
    with conn.cursor() as cur:
        returned = execute_values(
            cur,
            """
            INSERT INTO terminals (merchant_id, terminal_code, serial_number, model)
            VALUES %s
            RETURNING terminal_id, terminal_code
            """,
            [
                (
                    merchant_code_to_uuid[m["merchant_code"]],
                    m["terminal_code"],
                    m["serial_number"],
                    m["model"],
                )
                for m in terminal_meta
            ],
            page_size=200,
            fetch=True,
        )
        terminal_code_to_uuid = {r[1]: r[0] for r in returned}

    conn.commit()

    # Attach resolved UUIDs back to the info dicts
    for m in terminal_meta:
        m["merchant_id"] = merchant_code_to_uuid[m["merchant_code"]]
        m["terminal_id"] = terminal_code_to_uuid[m["terminal_code"]]

    return terminal_meta


# ──────────────────────────────────────────────────────────────────────────────
# INSERT CARDS
# ──────────────────────────────────────────────────────────────────────────────

def insert_cards(conn, count: int = 300) -> list[int]:
    """Generate synthetic masked cards. Returns list of card_ids."""
    rows = []
    for _ in range(count):
        brand     = random.choice(CARD_BRANDS)
        prefix    = random.choice(CARD_PREFIXES[brand])
        card_type = random.choices(CARD_TYPES, weights=[40, 50, 10])[0]
        bank      = random.choice(ISSUING_BANKS)
        exp_month = random.randint(1, 12)
        exp_year  = random.randint(2026, 2030)
        rows.append((
            mask_card_number(prefix),
            card_type,
            brand,
            bank,
            f"{exp_month:02d}/{exp_year}",
        ))

    with conn.cursor() as cur:
        execute_values(
            cur,
            """
            INSERT INTO cards (card_number_masked, card_type, card_brand, issuing_bank, expiry_date)
            VALUES %s
            RETURNING card_id
            """,
            rows,
            page_size=300,
        )
        card_ids = [r[0] for r in cur.fetchall()]
    conn.commit()
    return card_ids


# ──────────────────────────────────────────────────────────────────────────────
# VOLUME MULTIPLIER HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _get_holiday_multiplier(txn_date: date, category: str) -> float:
    """
    Return a volume multiplier for holiday/event surge periods.
    Only categories whose surge group appears in the period's dict are affected;
    all others return 1.0 (no change).
    """
    group = CATEGORY_SURGE_GROUP.get(category)
    if group is None:
        return 1.0
    for start, end, _label, mult_map in HOLIDAY_SURGE_PERIODS:
        if start <= txn_date <= end:
            return mult_map.get(group, 1.0)
    return 1.0


def _get_closure_multiplier(txn_date: date, closed_from: date | None) -> float:
    """
    Return a volume multiplier based on how many days until permanent closure.
    Returns 0.0 on or after the closure date (caller should skip the day).
    """
    if closed_from is None:
        return 1.0
    days_left = (closed_from - txn_date).days
    if days_left <= 0:
        return 0.0
    for threshold, factor in CLOSURE_DECLINE_STEPS:
        if days_left <= threshold:
            return factor
    return 1.0


# ──────────────────────────────────────────────────────────────────────────────
# TRANSACTION GENERATION
# ──────────────────────────────────────────────────────────────────────────────

def generate_and_insert_transactions(
    conn,
    merchants_info: list[dict],
    card_ids: list[int],
    qris_issuer_ids: dict[str, int],
    start_date: date | None = None,
    end_date: date | None = None,
) -> None:
    """
    For each merchant × each day in start_date..end_date (defaults to DATE_START..DATE_END):
      - Determine open window from schedule
      - Generate N transactions (volume per category)
      - ~85% SALE, ~3% REFUND, plus ~2% VOID companions for approved SALEs
      - Split EDC_CARD / QRIS per category probability
      - Insert in batches; build settlement + audit log records
    """
    qris_codes   = list(QRIS_ISSUER_WEIGHTS.keys())
    qris_weights = [QRIS_ISSUER_WEIGHTS[c] for c in qris_codes]
    qris_id_list = [qris_issuer_ids[c] for c in qris_codes]

    _start = start_date if start_date is not None else DATE_START
    _end   = end_date   if end_date   is not None else DATE_END

    date_range = [
        _start + timedelta(days=i)
        for i in range((_end - _start).days + 1)
    ]

    batch_counter = 1
    total_inserted = 0

    for info in merchants_info:
        category = info["category"]

        if category == "Bus Stop":
            continue  # no EDC terminal at bus stops

        m_id     = info["merchant_id"]
        t_id     = info["terminal_id"]
        schedule = info["schedule"]

        vol_min, vol_max = VOLUME_PROFILE.get(category, DEFAULT_VOLUME)
        amt_min, amt_max = AMOUNT_RANGE.get(category, DEFAULT_AMOUNT)
        qris_prob        = QRIS_PROBABILITY.get(category, DEFAULT_QRIS_PROB)
        closed_from      = info.get("closed_from")

        for txn_date in date_range:
            if closed_from and txn_date >= closed_from:
                continue  # permanently closed from this date

            weekday = txn_date.weekday()
            window  = schedule.get(weekday)
            if window is None:
                continue  # closed this weekday

            if txn_date in ID_NATIONAL_HOLIDAYS and category in HOLIDAY_CLOSED_CATEGORIES:
                continue  # closed for national holiday

            holiday_mult = _get_holiday_multiplier(txn_date, category)
            closure_mult = _get_closure_multiplier(txn_date, closed_from)

            open_t, close_t = window

            # For today: cap the close time to now so we never write future timestamps.
            if txn_date == date.today():
                now_local = datetime.now().time()
                if now_local <= open_t:
                    continue  # merchant hasn't opened yet
                if now_local < close_t:
                    close_t = now_local

            n_txns = max(1, int(random.randint(vol_min, vol_max) * holiday_mult * closure_mult))
            batch_number     = f"{batch_counter:03d}"
            batch_counter   += 1

            # Pre-sort timestamps to mimic sequential receipt flow
            timestamps = sorted(
                random_time_in_window(open_t, close_t, txn_date)
                for _ in range(n_txns)
            )

            txn_rows = []
            settle_sales_count   = 0
            settle_sales_amount  = Decimal("0")
            settle_refund_count  = 0
            settle_refund_amount = Decimal("0")

            for txn_time in timestamps:
                is_qris = random.random() < qris_prob
                if is_qris:
                    qris_id = random.choices(qris_id_list, weights=qris_weights)[0]
                    card_id = None
                    channel = "QRIS"
                else:
                    card_id = random.choice(card_ids)
                    qris_id = None
                    channel = "EDC_CARD"

                # QRIS never produces REFUND records — failed QRIS payments are
                # reversed automatically by the network without an EDC entry.
                tx_type = "SALE"
                if not is_qris and random.random() < 0.03:
                    tx_type = "REFUND"

                amount  = round_to_thousand(random.randint(amt_min, amt_max))

                if tx_type == "REFUND":
                    rc, status, appr = "00", "APPROVED", make_approval_code()
                else:
                    if is_qris:
                        rc = random.choices(QRIS_RESPONSE_CODE_POPULATION, weights=QRIS_RESPONSE_CODE_WEIGHTS)[0]
                    else:
                        rc = random.choices(RESPONSE_CODE_POPULATION, weights=RESPONSE_CODE_WEIGHTS)[0]
                    status = "APPROVED" if rc == "00" else "DECLINED"
                    appr   = make_approval_code() if rc == "00" else None

                settled = (rc == "00") and (txn_date < date.today())
                trace   = next_trace_number(txn_date)

                txn_rows.append((
                    trace, t_id, m_id,
                    card_id, qris_id, channel,
                    tx_type, float(amount), "IDR",
                    rc, RESPONSE_MESSAGES.get(rc, ""), appr,
                    status, batch_number, txn_time, settled,
                ))

                if tx_type == "SALE" and status == "APPROVED":
                    settle_sales_count  += 1
                    settle_sales_amount += amount
                elif tx_type == "REFUND" and status == "APPROVED":
                    settle_refund_count  += 1
                    settle_refund_amount += amount

            # Bulk-insert the day's transactions
            inserted_ids = _insert_transactions(conn, txn_rows)
            total_inserted += len(inserted_ids)

            # Generate VOID companions for ~2% of approved SALEs
            void_count = _insert_voids(conn, inserted_ids, m_id, t_id, batch_number, txn_date)
            total_inserted += void_count

            # Audit log (REQUEST + RESPONSE per transaction)
            _insert_logs(conn, inserted_ids)

            # Settlement summary for this merchant-day
            _upsert_settlement(
                conn, m_id, t_id, batch_number, txn_date,
                settle_sales_count,   settle_sales_amount,
                settle_refund_count,  settle_refund_amount,
            )

    print(f"      {total_inserted:,} transactions inserted.")


def _insert_transactions(conn, rows: list) -> list[int]:
    """Bulk-insert transaction rows; return list of transaction_ids."""
    if not rows:
        return []
    with conn.cursor() as cur:
        execute_values(
            cur,
            """
            INSERT INTO transactions
                (trace_number, terminal_id, merchant_id,
                 card_id, qris_issuer_id, payment_channel,
                 transaction_type, amount, currency,
                 response_code, response_message, approval_code,
                 transaction_status, batch_number, transaction_time, settled_flag)
            VALUES %s
            RETURNING transaction_id
            """,
            rows,
            page_size=500,
        )
        ids = [r[0] for r in cur.fetchall()]
    conn.commit()
    return ids


def _insert_voids(
    conn,
    txn_ids: list[int],
    merchant_id,
    terminal_id,
    batch_number: str,
    txn_date: date,
) -> int:
    """
    For ~2% of approved SALEs in txn_ids:
      1. Insert a VOID companion row (timestamped a few minutes later)
      2. Update the original SALE to REVERSED
    Returns the number of VOID rows inserted.
    """
    if not txn_ids:
        return 0

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT transaction_id, card_id, qris_issuer_id,
                   payment_channel, amount, transaction_time
            FROM   transactions
            WHERE  transaction_id    = ANY(%s)
              AND  transaction_type   = 'SALE'
              AND  transaction_status = 'APPROVED'
            """,
            (txn_ids,),
        )
        candidates = cur.fetchall()

    void_rows     = []
    reversed_ids  = []

    for txn_id, card_id, qris_id, channel, amount, txn_time in candidates:
        if random.random() > 0.02:
            continue
        void_time = txn_time + timedelta(minutes=random.randint(1, 5))
        void_rows.append((
            next_trace_number(txn_date),
            terminal_id, merchant_id,
            card_id, qris_id, channel,
            "VOID", float(amount), "IDR",
            "00", "APPROVED", make_approval_code(),
            "VOIDED", batch_number, void_time, False,
        ))
        reversed_ids.append(txn_id)

    if not void_rows:
        return 0

    with conn.cursor() as cur:
        execute_values(
            cur,
            """
            INSERT INTO transactions
                (trace_number, terminal_id, merchant_id,
                 card_id, qris_issuer_id, payment_channel,
                 transaction_type, amount, currency,
                 response_code, response_message, approval_code,
                 transaction_status, batch_number, transaction_time, settled_flag)
            VALUES %s
            """,
            void_rows,
            page_size=200,
        )
        cur.execute(
            """
            UPDATE transactions
            SET    transaction_status = 'REVERSED', settled_flag = FALSE
            WHERE  transaction_id = ANY(%s)
            """,
            (reversed_ids,),
        )
    conn.commit()
    return len(void_rows)


def _insert_logs(conn, txn_ids: list[int]) -> None:
    """Insert REQUEST + RESPONSE audit entries for the given transaction IDs."""
    if not txn_ids:
        return

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT transaction_id, trace_number, card_id, qris_issuer_id,
                   payment_channel, amount, response_code
            FROM   transactions
            WHERE  transaction_id = ANY(%s)
            """,
            (txn_ids,),
        )
        rows = cur.fetchall()

    log_rows = []
    for txn_id, trace, card_id, qris_id, channel, amount, rc in rows:
        ref = f'"pan_ref":"{card_id}"' if channel == "EDC_CARD" else f'"qris_ref":"{qris_id}"'
        req  = f'{{"mti":"0200","trace":"{trace}","amount":"{int(amount)}","channel":"{channel}",{ref}}}'
        resp = (
            f'{{"mti":"0210","trace":"{trace}","rc":"{rc}"'
            + (f',"approval":"{make_approval_code()}"' if rc == "00" else "")
            + "}}"
        )
        log_rows.append((txn_id, "REQUEST",  req))
        log_rows.append((txn_id, "RESPONSE", resp))

    with conn.cursor() as cur:
        execute_values(
            cur,
            "INSERT INTO transaction_log (transaction_id, log_type, raw_message) VALUES %s",
            log_rows,
            page_size=1000,
        )
    conn.commit()


def _upsert_settlement(
    conn,
    merchant_id,
    terminal_id,
    batch_number: str,
    settlement_date: date,
    sales_count: int,
    sales_amount: Decimal,
    refund_count: int,
    refund_amount: Decimal,
) -> None:
    net    = sales_amount - refund_amount
    status = "SETTLED" if settlement_date < date.today() else "PENDING"
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO settlement
                (merchant_id, terminal_id, batch_number, settlement_date,
                 total_sales_count, total_sales_amount,
                 total_refund_count, total_refund_amount,
                 net_amount, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (merchant_id, terminal_id, settlement_date)
            DO UPDATE SET
                total_sales_count   = EXCLUDED.total_sales_count,
                total_sales_amount  = EXCLUDED.total_sales_amount,
                total_refund_count  = EXCLUDED.total_refund_count,
                total_refund_amount = EXCLUDED.total_refund_amount,
                net_amount          = EXCLUDED.net_amount,
                status              = EXCLUDED.status
            """,
            (
                merchant_id, terminal_id, batch_number, settlement_date,
                sales_count,  float(sales_amount),
                refund_count, float(refund_amount),
                float(net),   status,
            ),
        )
    conn.commit()


# ──────────────────────────────────────────────────────────────────────────────
# LOAD FROM DB  (used by --purge to skip re-inserting reference rows)
# ──────────────────────────────────────────────────────────────────────────────

def load_merchants_from_db(conn, csv_rows: list[dict]) -> list[dict]:
    """
    Re-build the merchants_info list from existing DB rows + CSV schedule data.
    Matches by merchant_name so the correct operating hours are applied.
    """
    # Build a name → (category, schedule) lookup from the CSV
    csv_schedule: dict[str, dict] = {}
    for row in csv_rows:
        name     = row.get("name1", "").strip()
        category = row.get("primarycategorynm", "").strip()
        if not name:
            continue
        sched_override = row.get("_schedule_override")
        schedule = sched_override if sched_override else {wd: resolve_hours(row, wd, category) for wd in range(7)}
        cf_raw = row.get("closed_from", "")
        cf_str = cf_raw.strip() if isinstance(cf_raw, str) else ""
        bt      = row.get("batch_type", "").strip()
        xlsx_id = (row.get("_xlsx_id", "") or row.get("supplier_poiid", "")
                   or row.get("ID", "") or row.get("id", "")
                   or row.get("placeid", ""))
        csv_schedule[name] = {
            "category":    category,
            "schedule":    schedule,
            "closed_from": date.fromisoformat(cf_str) if cf_str else None,
            "batch_type":  bt if bt else None,
            "_xlsx_id":    xlsx_id if xlsx_id else None,
        }

    with conn.cursor() as cur:
        cur.execute("""
            SELECT m.merchant_id, m.merchant_name, m.merchant_code, m.mcc_code,
                   t.terminal_id, t.terminal_code, t.serial_number, t.model
            FROM   merchants m
            JOIN   terminals t ON t.merchant_id = m.merchant_id
        """)
        db_rows = cur.fetchall()

    # Reverse MCC lookup for fallback schedule when merchant isn't in csv_rows
    _mcc_to_cat = {v: k for k, v in CATEGORY_TO_MCC.items()}

    result = []
    for merchant_id, merchant_name, merchant_code, mcc_code, terminal_id, terminal_code, serial_number, model in db_rows:
        sched = csv_schedule.get(merchant_name)
        if sched is None:
            # Batch xlsx merchant not in csv_rows — build default schedule from MCC
            cat         = _mcc_to_cat.get(mcc_code, "")
            default_hrs = CATEGORY_DEFAULT_HOURS.get(cat)
            sched = {
                "category":    cat,
                "schedule":    {wd: default_hrs for wd in range(7)},
                "closed_from": None,
                "batch_type":  None,
            }
        result.append({
            "merchant_id":   merchant_id,
            "merchant_name": merchant_name,
            "merchant_code": merchant_code,
            "mcc_code":      mcc_code,
            "terminal_id":   terminal_id,
            "terminal_code": terminal_code,
            "serial_number": serial_number,
            "model":         model,
            "category":      sched.get("category", ""),
            "schedule":      sched.get("schedule", {}),
            "closed_from":   sched.get("closed_from"),
            "batch_type":    sched.get("batch_type"),
            "_xlsx_id":      sched.get("_xlsx_id"),
        })
    return result


def filter_new_csv_rows(conn, csv_rows: list[dict]) -> list[dict]:
    """Return only rows whose name1 is not already in the merchants table."""
    with conn.cursor() as cur:
        cur.execute("SELECT merchant_name FROM merchants")
        existing = {r[0] for r in cur.fetchall()}
    return [r for r in csv_rows if r.get("name1", "").strip() not in existing]


def get_merchant_idx_offset(conn) -> int:
    """Highest sequence number already embedded in merchant_code (5-digit suffix)."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT COALESCE(MAX(CAST(RIGHT(merchant_code, 5) AS INTEGER)), 0) FROM merchants"
        )
        return cur.fetchone()[0]


def load_cards_from_db(conn) -> list[int]:
    """Return existing card_ids from the DB."""
    with conn.cursor() as cur:
        cur.execute("SELECT card_id FROM cards")
        return [r[0] for r in cur.fetchall()]


def get_append_start_date(conn) -> date | None:
    """Return the day after the latest transaction date already in the DB, or None if empty."""
    with conn.cursor() as cur:
        cur.execute("SELECT MAX(transaction_time::date) FROM transactions")
        last_date = cur.fetchone()[0]
    return (last_date + timedelta(days=1)) if last_date else None


def reset_trace_seq(conn) -> None:
    """Advance the global trace-seq counter past all existing trace numbers to avoid collisions."""
    import helpers as _h_mod
    with conn.cursor() as cur:
        # Extract the full numeric suffix after 'RRN' + 8-char date (positions 1-11)
        cur.execute(
            "SELECT COALESCE(MAX(CAST(SUBSTRING(trace_number FROM 12) AS BIGINT)), 0) FROM transactions"
        )
        _h_mod._trace_seq = cur.fetchone()[0]


# ──────────────────────────────────────────────────────────────────────────────
# DATE-RANGE PURGE  (delete only the seeded date window, keep static tables)
# ──────────────────────────────────────────────────────────────────────────────

def purge_date_range(conn) -> None:
    """
    Delete all transaction data for DATE_START..DATE_END without touching
    merchants, terminals, cards, admin_areas, or reference tables.
    Useful for re-seeding a date range without a full reset.
    """
    start_ts = datetime.combine(DATE_START, time.min)
    end_ts   = datetime.combine(DATE_END,   time.max)

    with conn.cursor() as cur:
        cur.execute("""
            DELETE FROM transaction_log
            WHERE transaction_id IN (
                SELECT transaction_id FROM transactions
                WHERE transaction_time BETWEEN %s AND %s
            )
        """, (start_ts, end_ts))
        logs_deleted = cur.rowcount

        cur.execute("""
            DELETE FROM transactions
            WHERE transaction_time BETWEEN %s AND %s
        """, (start_ts, end_ts))
        txn_deleted = cur.rowcount

        cur.execute("""
            DELETE FROM settlement
            WHERE settlement_date BETWEEN %s AND %s
        """, (DATE_START, DATE_END))
        settle_deleted = cur.rowcount

    conn.commit()
    print(f"      Purged {txn_deleted:,} transactions, "
          f"{logs_deleted:,} log entries, "
          f"{settle_deleted:,} settlement rows "
          f"({DATE_START} – {DATE_END}).")


# ──────────────────────────────────────────────────────────────────────────────
# PRUNE CLOSED MERCHANTS  (remove data on/after each merchant's closed_from date)
# ──────────────────────────────────────────────────────────────────────────────

def prune_closed_merchants(conn, csv_rows: list[dict]) -> None:
    """
    For every merchant in the CSV that has a non-empty closed_from date,
    delete all transaction_log / transactions / settlement rows on or after
    that date.  Merchant, terminal, card, and admin-area rows are untouched.
    """
    closure_map: dict[str, date] = {}
    for row in csv_rows:
        name   = row.get("name1", "").strip()
        cf_str = row.get("closed_from", "").strip()
        if name and cf_str:
            try:
                closure_map[name] = date.fromisoformat(cf_str)
            except ValueError:
                print(f"      WARNING: invalid closed_from '{cf_str}' for '{name}' — skipped.")

    if not closure_map:
        print("      No closed_from dates found in CSV — nothing to prune.")
        return

    with conn.cursor() as cur:
        cur.execute("SELECT merchant_name, merchant_id FROM merchants")
        db_map = {name: mid for name, mid in cur.fetchall()}

    total_txn = total_log = total_settle = 0

    for name, closed_from in sorted(closure_map.items(), key=lambda x: x[1]):
        m_id = db_map.get(name)
        if m_id is None:
            print(f"      WARNING: '{name}' not found in DB — skipped.")
            continue

        close_ts = datetime.combine(closed_from, time.min)
        with conn.cursor() as cur:
            cur.execute("""
                DELETE FROM transaction_log
                WHERE transaction_id IN (
                    SELECT transaction_id FROM transactions
                    WHERE  merchant_id      = %s
                      AND  transaction_time >= %s
                )
            """, (m_id, close_ts))
            logs_del = cur.rowcount

            cur.execute("""
                DELETE FROM transactions
                WHERE merchant_id      = %s
                  AND transaction_time >= %s
            """, (m_id, close_ts))
            txn_del = cur.rowcount

            cur.execute("""
                DELETE FROM settlement
                WHERE merchant_id    = %s
                  AND settlement_date >= %s
            """, (m_id, closed_from))
            settle_del = cur.rowcount

        conn.commit()
        total_txn    += txn_del
        total_log    += logs_del
        total_settle += settle_del

        print(f"      {name[:38]:<38}  closed {closed_from}"
              f"  → -{txn_del:,} txn  -{logs_del:,} log  -{settle_del:,} settle")

    print(f"\n      Total: -{total_txn:,} transactions, "
          f"-{total_log:,} log entries, "
          f"-{total_settle:,} settlement rows.")


# ──────────────────────────────────────────────────────────────────────────────
# MERCHANT STATUS REPORT
# ──────────────────────────────────────────────────────────────────────────────

def _save_card_jpeg(
    merchant_name: str,
    status: str,
    confidence: float,
    last_txn_fmt: str,
    ago_str: str,
    c24h: int, c7d: int, c30d: int,
    qris_30d: int, edc_30d: int,
    active_days: int, window_days: int,
    activity_ratio: float, max_gap: int,
    reasons: list[str],
    now_wib,
    output_dir: Path,
    seq: int,
    xlsx_id: str = "",
) -> "Path | None":
    """Render a styled JPEG activity card for one merchant. Returns the saved path."""
    try:
        from PIL import Image, ImageDraw
    except ImportError:
        return None

    font_sm  = _load_font(13)
    font_md  = _load_font(15)
    font_lg  = _load_font(18)

    status_colour = _C_ACTIVE if status == "ACTIVE" else _C_INACTIVE

    # ── compute card height dynamically ───────────────────────────────────────
    n_lines  = 6                     # name + 5 data rows
    n_reason = len(reasons)
    height   = _PAD + _LH + 6 + _LH + (_LH * n_lines) + 10 + (_LH * n_reason) + _PAD + 20

    img  = Image.new("RGB", (_CARD_W, max(height, 240)), _C_BG)
    draw = ImageDraw.Draw(img)

    # outer panel
    draw.rounded_rectangle([8, 8, _CARD_W - 8, height - 8], radius=10,
                            fill=_C_PANEL, outline=_C_BORDER, width=1)

    y = _PAD

    # merchant name
    draw.text((_PAD, y), merchant_name, font=font_lg, fill=_C_NAME)
    y += _LH + 6

    # thin divider
    draw.line([_PAD, y, _CARD_W - _PAD, y], fill=_C_DIVIDER, width=1)
    y += 8

    def row(label: str, value: str, value_colour: str = _C_VALUE) -> None:
        nonlocal y
        draw.text((_PAD, y), f"{label:<14}", font=font_md, fill=_C_LABEL)
        draw.text((_PAD + 14 * 9, y), value, font=font_md, fill=value_colour)
        y += _LH

    row("STATUS",        f"{status}   (confidence {confidence:.3f})", status_colour)
    row("last_txn",      f"{last_txn_fmt}  ({ago_str})")
    row("24h / 7d / 30d", f"{c24h}  /  {c7d}  /  {c30d}")
    row("channel split", f"QRIS {qris_30d}   |   EDC {edc_30d}")
    row("active days",   f"{active_days}/{window_days}  (ratio {activity_ratio}, max gap {max_gap}d)")

    if reasons:
        y += 4
        draw.line([_PAD, y, _CARD_W - _PAD, y], fill=_C_DIVIDER, width=1)
        y += 8
        for r in reasons:
            draw.text((_PAD, y), f"  -  {r}", font=font_sm, fill=_C_REASON)
            y += _LH

    # footer timestamp
    y += 6
    footer = f"Generated {now_wib.strftime('%Y-%m-%dT%H:%M:%S%z')}"
    draw.text((_PAD, y), footer, font=font_sm, fill=_C_BORDER)

    # save — use xlsx_id for filename when available
    if xlsx_id:
        safe_id = "".join(c if c.isalnum() or c in "-_" else "_" for c in xlsx_id)
        fname = output_dir / f"{safe_id[:60]}.jpg"
    else:
        safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in merchant_name)
        fname = output_dir / f"{seq:02d}_{safe_name[:40].strip()}.jpg"
    img.save(fname, "JPEG", quality=92)
    return fname


def _save_realtime_card_jpeg(
    name: str,
    status: str,
    last_signal_fmt: str,
    ago_str: str,
    reasons: list,
    now_wib,
    output_dir: Path,
    seq: int,
    xlsx_id: str = "",
) -> "Path | None":
    """Render a simplified signal-only card for realtime (ATM/bank) POIs."""
    try:
        from PIL import Image, ImageDraw
    except ImportError:
        return None

    font_sm = _load_font(13)
    font_md = _load_font(15)
    font_lg = _load_font(18)
    status_colour = _C_ACTIVE if status == "ACTIVE" else _C_INACTIVE

    n_reason = len(reasons)
    height = _PAD + _LH + 6 + _LH + (_LH * 2) + 10 + (_LH * n_reason) + _PAD + 20

    img  = Image.new("RGB", (_CARD_W, max(height, 180)), _C_BG)
    draw = ImageDraw.Draw(img)
    draw.rounded_rectangle([8, 8, _CARD_W - 8, height - 8], radius=10,
                            fill=_C_PANEL, outline=_C_BORDER, width=1)
    y = _PAD
    draw.text((_PAD, y), name, font=font_lg, fill=_C_NAME)
    y += _LH + 6
    draw.line([_PAD, y, _CARD_W - _PAD, y], fill=_C_DIVIDER, width=1)
    y += 8

    def _row(label, value, vc=_C_VALUE):
        nonlocal y
        draw.text((_PAD, y), f"{label:<14}", font=font_md, fill=_C_LABEL)
        draw.text((_PAD + 14 * 9, y), value, font=font_md, fill=vc)
        y += _LH

    _row("STATUS",      status, status_colour)
    _row("last_signal", f"{last_signal_fmt}  ({ago_str})")

    if reasons:
        y += 4
        draw.line([_PAD, y, _CARD_W - _PAD, y], fill=_C_DIVIDER, width=1)
        y += 8
        for r in reasons:
            draw.text((_PAD, y), f"  -  {r}", font=font_sm, fill=_C_REASON)
            y += _LH

    y += 6
    draw.text((_PAD, y), f"Generated {now_wib.strftime('%Y-%m-%dT%H:%M:%S%z')}",
              font=font_sm, fill=_C_BORDER)

    if xlsx_id:
        safe_id = "".join(c if c.isalnum() or c in "-_" else "_" for c in xlsx_id)
        fname = output_dir / f"{safe_id[:60]}_rt.jpg"
    else:
        safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in name)
        fname = output_dir / f"{seq:02d}_{safe_name[:40].strip()}_rt.jpg"
    img.save(fname, "JPEG", quality=92)
    return fname


def _save_excel_report(
    records: list[tuple[int, str, str, "Path | None"]],
    output_dir: Path,
    now_wib,
) -> "Path | None":
    """
    Write merchant_activity_report.xlsx to output_dir.
    records = [(seq, poi_name, last_signal_str, jpeg_path), ...]
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("      [excel] openpyxl not installed — skipping Excel export")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Merchant Activity"

    HDR_FILL = PatternFill("solid", fgColor="1E2A3A")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers    = ["#",  "POI Name", "Last Signal Update",  "Proof (IMG)"]
    col_widths = [4,    36,         28,                     48          ]

    ws.row_dimensions[1].height = 28
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = HDR_ALIGN
        ws.column_dimensions[cell.column_letter].width = w

    IMG_W  = 320   # pixels in Excel
    IMG_H  = 160   # pixels in Excel
    ROW_H  = 122   # row height in points (≈ 163px @ 96 dpi)

    CTR = Alignment(horizontal="center", vertical="center")
    MID = Alignment(vertical="center", wrap_text=True)

    for row_idx, (seq_n, name, last_signal, jpeg_path) in enumerate(records, start=2):
        ws.row_dimensions[row_idx].height = ROW_H
        ws.cell(row=row_idx, column=1, value=seq_n).alignment    = CTR
        ws.cell(row=row_idx, column=2, value=name).alignment     = MID
        ws.cell(row=row_idx, column=3, value=last_signal).alignment = CTR
        if jpeg_path and Path(jpeg_path).exists():
            xl_img        = XLImage(str(jpeg_path))
            xl_img.width  = IMG_W
            xl_img.height = IMG_H
            ws.add_image(xl_img, f"D{row_idx}")

    xlsx_path = output_dir / "merchant_activity_report.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def _save_batch_excel_report(
    new_records:      list,
    update_records:   list,
    delete_records:   list,
    realtime_records: list,
    output_dir: Path,
    now_wib,
) -> "Path | None":
    """Four-sheet Excel (NEW / UPDATE / DELETE / REALTIME) for batch merchants."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return None

    wb = Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1E2A3A")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    headers    = ["#", "ID", "POI Name", "Last Signal Update", "Proof (IMG)"]
    col_widths = [4, 36, 36, 28, 48]
    IMG_W, IMG_H, ROW_H = 320, 160, 122

    def _write_sheet(ws, records, sheet_title):
        ws.title = sheet_title
        ws.row_dimensions[1].height = 28
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN
            ws.column_dimensions[cell.column_letter].width = w
        CTR = Alignment(horizontal="center", vertical="center")
        MID = Alignment(vertical="center", wrap_text=True)
        for row_idx, (seq_n, name, last_signal, jpeg_path, xlsx_id) in enumerate(records, start=2):
            ws.row_dimensions[row_idx].height = ROW_H
            ws.cell(row=row_idx, column=1, value=seq_n).alignment    = CTR
            ws.cell(row=row_idx, column=2, value=xlsx_id).alignment  = MID
            ws.cell(row=row_idx, column=3, value=name).alignment     = MID
            ws.cell(row=row_idx, column=4, value=last_signal).alignment = CTR
            if jpeg_path and Path(jpeg_path).exists():
                xl_img        = XLImage(str(jpeg_path))
                xl_img.width  = IMG_W
                xl_img.height = IMG_H
                ws.add_image(xl_img, f"E{row_idx}")

    ws_new = wb.active
    _write_sheet(ws_new, new_records, "NEW")
    ws_upd = wb.create_sheet()
    _write_sheet(ws_upd, update_records, "UPDATE")
    ws_del = wb.create_sheet()
    _write_sheet(ws_del, delete_records, "DELETE")
    ws_rt = wb.create_sheet()
    _write_sheet(ws_rt, realtime_records, "REALTIME")

    xlsx_path = output_dir / "batch_activity_report.xlsx"
    wb.save(xlsx_path)

    # Write companion meta JSON so --upload can build the Drive-linked version
    import json as _json
    _meta = {}
    for sheet_name, recs in (
        ("NEW", new_records), ("UPDATE", update_records),
        ("DELETE", delete_records), ("REALTIME", realtime_records),
    ):
        _meta[sheet_name] = [
            {"seq": s, "name": n, "last_signal": ls,
             "jpeg_path": str(jp) if jp else "", "xlsx_id": xid}
            for s, n, ls, jp, xid in recs
        ]
    (output_dir / "batch_activity_report_meta.json").write_text(
        _json.dumps(_meta, indent=2, ensure_ascii=False)
    )

    return xlsx_path


def _build_linked_batch_excel(meta: dict, drive_links: dict, output_dir: Path) -> Path:
    """
    Build batch_activity_report.xlsx with Drive hyperlinks in the Proof column
    instead of embedded images. Used for the version uploaded to Google Drive.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    HDR_FILL  = PatternFill("solid", fgColor="1E2A3A")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LINK_FONT = Font(color="0563C1", underline="single")
    headers    = ["#", "ID", "POI Name", "Last Signal Update", "Proof (Link)"]
    col_widths = [4, 36, 36, 28, 36]

    def _write_linked(ws, records, title):
        ws.title = title
        ws.row_dimensions[1].height = 28
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN
            ws.column_dimensions[cell.column_letter].width = w
        CTR = Alignment(horizontal="center", vertical="center")
        MID = Alignment(vertical="center", wrap_text=True)
        for ri, rec in enumerate(records, start=2):
            ws.row_dimensions[ri].height = 18
            ws.cell(row=ri, column=1, value=rec["seq"]).alignment     = CTR
            ws.cell(row=ri, column=2, value=rec["xlsx_id"]).alignment = MID
            ws.cell(row=ri, column=3, value=rec["name"]).alignment    = MID
            ws.cell(row=ri, column=4, value=rec["last_signal"]).alignment = CTR
            url = drive_links.get(rec.get("jpeg_path", ""), "")
            if url:
                c = ws.cell(row=ri, column=5, value="view")
                c.hyperlink = url
                c.font = LINK_FONT
                c.alignment = CTR
            else:
                ws.cell(row=ri, column=5, value="—").alignment = CTR

    _write_linked(wb.active, meta.get("NEW", []), "NEW")
    for sheet_name in ("UPDATE", "DELETE", "REALTIME"):
        _write_linked(wb.create_sheet(), meta.get(sheet_name, []), sheet_name)

    out = output_dir / "batch_activity_report_drive.xlsx"
    wb.save(str(out))
    return out


def _generate_realtime_records(output_dir: Path, seq_start: int, now_wib) -> "tuple[list, int]":
    """Read list_realtime.csv, generate signal-only JPEG cards, return (records, next_seq)."""
    import csv as _csv_mod
    records = []
    seq = seq_start
    if not _REALTIME_CSV.exists():
        return records, seq

    with open(_REALTIME_CSV, newline="", encoding="utf-8") as f:
        for row in _csv_mod.DictReader(f):
            raw_name  = row.get("poi_nm", "")
            street    = row.get("street", "")
            clean_st  = re.sub(r"(?i)^jalan\s+", "", street).strip()
            name      = f"{raw_name.upper()} {clean_st}".strip() if clean_st else raw_name.upper()
            xlsx_id   = row.get("id", "")
            sig_raw   = (row.get("last_signal") or "").strip()

            sig_dt = None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    sig_dt = datetime.strptime(sig_raw, fmt).replace(tzinfo=WIB)
                    break
                except ValueError:
                    pass

            if sig_dt:
                delta_h  = (now_wib - sig_dt).total_seconds() / 3600
                ago_str  = _format_ago(delta_h)
                days_ago = delta_h / 24
                status   = "ACTIVE" if days_ago <= REALTIME_ACTIVE_DAYS else "INACTIVE"
                sig_fmt  = f"~{sig_dt.strftime('%Y-%m-%d')}"
                reasons  = [] if status == "ACTIVE" else [f"last signal {ago_str}"]
            else:
                ago_str = "unknown"
                status  = "INACTIVE"
                sig_fmt = sig_raw or "—"
                reasons = ["signal date unavailable"]

            jpeg_path = _save_realtime_card_jpeg(
                name=name, status=status,
                last_signal_fmt=sig_fmt, ago_str=ago_str,
                reasons=reasons, now_wib=now_wib,
                output_dir=output_dir, seq=seq,
                xlsx_id=xlsx_id,
            )
            records.append((seq, name, f"{sig_fmt}  ({ago_str})", jpeg_path, xlsx_id))
            seq += 1

    return records, seq


def generate_merchant_status_report(conn, merchants_info: list[dict]) -> None:
    """
    Print a live-signal activity health card for every seeded merchant,
    modelled on the card shown in the product spec.

    Reference "now" is end-of-day on DATE_END (the last seeded day).
    Windows:
      24h  = last 24 h before reference time
       7d  = DATE_END minus 6 days  (7 calendar days inclusive)
      30d  = DATE_END minus 29 days (30 calendar days inclusive)
    """
    now_wib        = datetime.now(WIB)
    end_of_data    = datetime.combine(DATE_END, time(23, 59, 59)).replace(tzinfo=WIB)
    # Use end-of-DATE_END when seeded data extends past the current clock
    # (transactions near closing time today would otherwise show negative hours).
    ref_time       = max(now_wib, end_of_data)
    window_30d     = (DATE_END - timedelta(days=29))
    window_7d      = (DATE_END - timedelta(days=6))
    window_24h     = datetime.combine(DATE_END, time(23, 59, 59)).replace(tzinfo=WIB) - timedelta(hours=24)
    _WINDOW_DAYS   = WINDOW_DAYS

    output_dir = Path(__file__).parent / "reports" / DATE_END.strftime("%Y%m%d")
    output_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 62)
    print("  MERCHANT ACTIVITY REPORT")
    print(f"  Loaded at : {now_wib.strftime('%Y-%m-%dT%H:%M:%S%z')}  (WIB UTC+7)")
    print(f"  Data range: {DATE_START} – {DATE_END}")
    print(f"  Output    : {output_dir}")
    print("=" * 62)

    seq = 1
    excel_records: list[tuple[int, str, str, "Path | None"]] = []
    batch_records: list[tuple[int, str, str, "Path | None", str, str]] = []

    with conn.cursor() as cur:
        for info in merchants_info:

            m_id   = info["merchant_id"]
            m_name = info["merchant_name"]

            # ── last approved transaction ──────────────────────────────────
            cur.execute("""
                SELECT MAX(transaction_time)
                FROM transactions
                WHERE merchant_id = %s AND transaction_status = 'APPROVED'
            """, (m_id,))
            last_txn_utc = cur.fetchone()[0]

            if last_txn_utc is None:
                print(f"\n{'─'*62}\n{m_name}")
                closed_from_val = info.get("closed_from")
                if closed_from_val:
                    # Closed before the seeded period — synthesise ago from the closure date
                    h_ago = (ref_time - datetime.combine(closed_from_val, time(23, 59, 59)).replace(tzinfo=WIB)).total_seconds() / 3600
                    ago_s = _format_ago(h_ago)
                    ltf   = f"~{closed_from_val}"
                    print(f"STATUS        : {'INACTIVE':<10} (permanently closed)")
                    print(f"last_txn      : {ltf}  ({ago_s})")
                    bt      = info.get("batch_type")
                    xlsx_id = info.get("_xlsx_id") or ""
                    jpeg_path = _save_card_jpeg(
                        merchant_name=m_name, status="INACTIVE", confidence=0.0,
                        last_txn_fmt=ltf, ago_str=ago_s,
                        c24h=0, c7d=0, c30d=0, qris_30d=0, edc_30d=0,
                        active_days=0, window_days=_WINDOW_DAYS,
                        activity_ratio=0.0, max_gap=0,
                        reasons=[f"last txn {ago_s}"],
                        now_wib=now_wib, output_dir=output_dir, seq=seq,
                        xlsx_id=xlsx_id,
                    )
                    excel_records.append((seq, m_name, f"{ltf}  ({ago_s})", jpeg_path))
                    if bt in ("xlsx-new", "xlsx-update", "xlsx-delete"):
                        batch_records.append((seq, m_name, f"{ltf}  ({ago_s})", jpeg_path, bt, xlsx_id))
                    seq += 1
                else:
                    print("STATUS        : NO DATA")
                    bt      = info.get("batch_type")
                    xlsx_id = info.get("_xlsx_id") or ""
                    if bt in ("xlsx-new", "xlsx-update", "xlsx-delete"):
                        jpeg_path = _save_card_jpeg(
                            merchant_name=m_name, status="INACTIVE", confidence=0.0,
                            last_txn_fmt="—", ago_str="no data",
                            c24h=0, c7d=0, c30d=0, qris_30d=0, edc_30d=0,
                            active_days=0, window_days=_WINDOW_DAYS,
                            activity_ratio=0.0, max_gap=0,
                            reasons=["no transaction data found"],
                            now_wib=now_wib, output_dir=output_dir, seq=seq,
                            xlsx_id=xlsx_id,
                        )
                        excel_records.append((seq, m_name, "—  (no data)", jpeg_path))
                        batch_records.append((seq, m_name, "—  (no data)", jpeg_path, bt, xlsx_id))
                        seq += 1
                continue

            last_txn_wib = last_txn_utc.astimezone(WIB)
            hours_ago    = (ref_time - last_txn_wib).total_seconds() / 3600

            # ── transaction counts (approved SALE only) ────────────────────
            cur.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE transaction_time >= %s)                             AS c24h,
                    COUNT(*) FILTER (WHERE transaction_time >= %s::date)                       AS c7d,
                    COUNT(*) FILTER (WHERE transaction_time >= %s::date)                       AS c30d,
                    COUNT(*) FILTER (WHERE transaction_time >= %s::date
                                      AND payment_channel = 'QRIS')                           AS qris,
                    COUNT(*) FILTER (WHERE transaction_time >= %s::date
                                      AND payment_channel = 'EDC_CARD')                       AS edc
                FROM transactions
                WHERE merchant_id        = %s
                  AND transaction_status = 'APPROVED'
                  AND transaction_type   = 'SALE'
            """, (window_24h, window_7d, window_30d, window_30d, window_30d, m_id))
            c24h, c7d, c30d, qris_30d, edc_30d = cur.fetchone()

            # ── active days + max gap ──────────────────────────────────────
            cur.execute("""
                SELECT DISTINCT transaction_time::date AS d
                FROM   transactions
                WHERE  merchant_id        = %s
                  AND  transaction_status = 'APPROVED'
                  AND  transaction_time  >= %s::date
                ORDER  BY 1
            """, (m_id, window_30d))
            active_dates = [r[0] for r in cur.fetchall()]
            active_days  = len(active_dates)

            max_gap = 1
            if len(active_dates) >= 2:
                gaps    = [(active_dates[i+1] - active_dates[i]).days for i in range(len(active_dates) - 1)]
                max_gap = max(gaps)

            activity_ratio = round(active_days / _WINDOW_DAYS, 3)

            # ── confidence score ───────────────────────────────────────────
            recency_score = (
                1.0 if hours_ago <=  24 else
                0.9 if hours_ago <=  72 else
                0.7 if hours_ago <= 168 else
                0.3
            )
            activity_score = min(active_days / _WINDOW_DAYS, 1.0)
            volume_score   = min(math.log10(max(c30d, 1)) / math.log10(500), 1.0)
            confidence     = round(recency_score * 0.40 + activity_score * 0.35 + volume_score * 0.25, 3)

            # ── status (ACTIVE = at least one txn within 90 days) ─────────
            status = "ACTIVE" if hours_ago <= 90 * 24 else "INACTIVE"

            # ── format helpers ─────────────────────────────────────────────
            ago_str      = _format_ago(hours_ago)
            last_txn_fmt = last_txn_wib.strftime("%Y-%m-%dT%H:%M:%S%z")

            # ── reason bullets ─────────────────────────────────────────────
            reasons: list[str] = []
            if hours_ago <= 72:
                reasons.append(f"last txn {int(hours_ago)}h ago (<= 72h)")
            if active_days >= _WINDOW_DAYS * 0.8:
                reasons.append(f"{active_days}/{_WINDOW_DAYS} active days, max gap {max_gap}d")
            if c30d >= 50:
                reasons.append(f"{c30d} approved txn / 30d ({qris_30d} QRIS, {edc_30d} EDC)")

            # ── override: permanently closed merchants ─────────────────────
            closed_from_val = info.get("closed_from")
            if closed_from_val:
                status = "INACTIVE"
                reasons.insert(0, f"last txn {ago_str}")

            # ── print card ─────────────────────────────────────────────────
            print(f"\n{'─'*62}")
            print(f"{m_name}")
            print(f"STATUS        : {status:<10} (confidence {confidence:.3f})")
            print(f"last_txn      : {last_txn_fmt}  ({ago_str})")
            print(f"txn 24h/7d/30d: {c24h} / {c7d} / {c30d}")
            print(f"channel split : QRIS {qris_30d}  |  EDC {edc_30d}")
            print(f"active days   : {active_days}/{_WINDOW_DAYS}  (ratio {activity_ratio}, max gap {max_gap}d)")
            if reasons:
                print("reasons       :")
                for r in reasons:
                    print(f"  - {r}")

            # ── save JPEG card ─────────────────────────────────────────────
            bt      = info.get("batch_type")
            xlsx_id = info.get("_xlsx_id") or ""
            jpeg_path = _save_card_jpeg(
                merchant_name=m_name,
                status=status,
                confidence=confidence,
                last_txn_fmt=last_txn_fmt,
                ago_str=ago_str,
                c24h=c24h, c7d=c7d, c30d=c30d,
                qris_30d=qris_30d, edc_30d=edc_30d,
                active_days=active_days, window_days=_WINDOW_DAYS,
                activity_ratio=activity_ratio, max_gap=max_gap,
                reasons=reasons,
                now_wib=now_wib,
                output_dir=output_dir,
                seq=seq,
                xlsx_id=xlsx_id,
            )
            excel_records.append((seq, m_name, f"{last_txn_fmt}  ({ago_str})", jpeg_path))
            if bt in ("xlsx-new", "xlsx-update", "xlsx-delete"):
                batch_records.append((seq, m_name, f"{last_txn_fmt}  ({ago_str})", jpeg_path, bt, xlsx_id))
            seq += 1

    xlsx_path = _save_excel_report(excel_records, output_dir, now_wib)

    batch_new    = [(s, n, ls, jp, xid) for s, n, ls, jp, bt, xid in batch_records if bt == "xlsx-new"]
    batch_update = [(s, n, ls, jp, xid) for s, n, ls, jp, bt, xid in batch_records if bt == "xlsx-update"]
    batch_delete = [(s, n, ls, jp, xid) for s, n, ls, jp, bt, xid in batch_records if bt == "xlsx-delete"]
    realtime_records, seq = _generate_realtime_records(output_dir, seq, now_wib)
    if batch_new or batch_update or batch_delete or realtime_records:
        batch_path = _save_batch_excel_report(
            batch_new, batch_update, batch_delete, realtime_records, output_dir, now_wib
        )
    else:
        batch_path = None

    print(f"\n{'='*62}")
    print(f"  JPEG cards saved to: {output_dir}")
    if xlsx_path:
        print(f"  Excel report     : {xlsx_path}")
    if batch_path:
        print(f"  Batch report     : {batch_path}")
    print()


# ──────────────────────────────────────────────────────────────────────────────
# GOOGLE DRIVE UPLOAD
# ──────────────────────────────────────────────────────────────────────────────

def _get_gdrive_service():
    """Return an authenticated Google Drive API v3 service object."""
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    creds = None
    if GDRIVE_TOKEN.exists():
        creds = Credentials.from_authorized_user_file(str(GDRIVE_TOKEN), GDRIVE_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(GDRIVE_CREDENTIALS), GDRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        GDRIVE_TOKEN.write_text(creds.to_json())
    return build("drive", "v3", credentials=creds)


def _upload_to_gdrive(output_dir: Path) -> None:
    """Upload Excel reports from output_dir into a dated subfolder on Google Drive."""
    from googleapiclient.http import MediaFileUpload

    if not GDRIVE_FOLDER_ID:
        print("  [gdrive] GDRIVE_FOLDER_ID not set — skipping upload.")
        return
    if not GDRIVE_CREDENTIALS.exists():
        print(f"  [gdrive] credentials.json not found at {GDRIVE_CREDENTIALS} — skipping.")
        return

    service = _get_gdrive_service()
    date_folder = output_dir.name   # e.g. "20260624"

    # Find or create a dated subfolder inside the designated parent folder
    q = (f"name='{date_folder}' and "
         f"mimeType='application/vnd.google-apps.folder' and "
         f"'{GDRIVE_FOLDER_ID}' in parents and trashed=false")
    res = service.files().list(q=q, fields="files(id)").execute()
    if res["files"]:
        sub_id = res["files"][0]["id"]
    else:
        meta = {
            "name": date_folder,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [GDRIVE_FOLDER_ID],
        }
        sub_id = service.files().create(body=meta, fields="id").execute()["id"]
        print(f"  [gdrive] Created folder : {date_folder}")

    def _upsert(fname: str, local_path: Path, mime: str) -> str:
        """Upload or update a file in sub_id; return the Drive file ID."""
        media = MediaFileUpload(str(local_path), mimetype=mime, resumable=True)
        q2 = f"name='{fname}' and '{sub_id}' in parents and trashed=false"
        existing = service.files().list(q=q2, fields="files(id)").execute()["files"]
        if existing:
            fid = existing[0]["id"]
            service.files().update(fileId=fid, media_body=media).execute()
            print(f"  [gdrive] Updated  : {fname}")
        else:
            fid = service.files().create(
                body={"name": fname, "parents": [sub_id]},
                media_body=media, fields="id",
            ).execute()["id"]
            print(f"  [gdrive] Uploaded : {fname}")
        return fid

    # Upload merchant_activity_report.xlsx (unchanged, embedded images)
    xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if (output_dir / "merchant_activity_report.xlsx").exists():
        _upsert("merchant_activity_report.xlsx",
                output_dir / "merchant_activity_report.xlsx", xlsx_mime)

    # If meta JSON exists: upload JPEGs → build Drive-linked batch xlsx → upload it
    import json as _json
    meta_path = output_dir / "batch_activity_report_meta.json"
    if meta_path.exists() and (output_dir / "batch_activity_report.xlsx").exists():
        meta = _json.loads(meta_path.read_text())
        drive_links: dict[str, str] = {}
        jpeg_count = 0
        for recs in meta.values():
            for rec in recs:
                jp = rec.get("jpeg_path", "")
                if not jp or not Path(jp).exists():
                    continue
                fid = _upsert(Path(jp).name, Path(jp), "image/jpeg")
                drive_links[jp] = f"https://drive.google.com/file/d/{fid}/view"
                jpeg_count += 1
        print(f"  [gdrive] JPEGs    : {jpeg_count} uploaded")
        linked_path = _build_linked_batch_excel(meta, drive_links, output_dir)
        _upsert("batch_activity_report.xlsx", linked_path, xlsx_mime)
    elif (output_dir / "batch_activity_report.xlsx").exists():
        _upsert("batch_activity_report.xlsx",
                output_dir / "batch_activity_report.xlsx", xlsx_mime)


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def _load_xlsx_supplement(csv_rows: list[dict]) -> int:
    """
    Append list_edc.csv rows to csv_rows in-place (for schedule lookup / batch tagging).
    Returns the number of rows appended.
    """
    count = 0
    if _BATCH_UNIFIED_CSV.exists():
        rows_u = _read_unified_csv(_BATCH_UNIFIED_CSV)
        non_rt = [r for r in rows_u if r.get("batch_type") != "xlsx-realtime"]
        csv_rows.extend(non_rt)
        count += len(non_rt)
    return count


def main() -> None:
    # Ensure UTF-8 output on Windows (─ and ═ chars in report cards)
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    purge_mode           = "--purge"           in sys.argv
    report_mode          = "--report"          in sys.argv
    report_selected_mode = "--report-selected" in sys.argv
    prune_closed_mode    = "--prune-closed"    in sys.argv
    append_mode          = "--append"          in sys.argv
    reset_mode           = "--reset"           in sys.argv
    add_merchants_mode   = "--add-merchants"   in sys.argv
    batch_seed_mode      = "--batch-seed"      in sys.argv
    upload_mode          = "--upload"          in sys.argv

    conn = get_connection()

    if upload_mode:
        reports_root = Path(__file__).parent / "reports"
        # Pick the most recently modified YYYYMMDD subfolder
        candidates = sorted(
            [p for p in reports_root.iterdir() if p.is_dir() and p.name.isdigit()],
            key=lambda p: p.stat().st_mtime, reverse=True,
        ) if reports_root.exists() else []
        if not candidates:
            print(f"  No report folders found under {reports_root}")
            print(f"  Run --report or --batch-seed first to generate the Excel files.")
            conn.close()
            return
        output_dir = candidates[0]
        print(f"[upload] Uploading Excel reports from {output_dir} ...")
        _upload_to_gdrive(output_dir)
        conn.close()
        return

    # Auto-detect: if no explicit mode and DB already has data → append instead of wipe
    if not any([reset_mode, purge_mode, append_mode, report_mode, report_selected_mode,
                prune_closed_mode, add_merchants_mode, batch_seed_mode]):
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM transactions")
            if cur.fetchone()[0] > 0:
                append_mode = True

    if report_selected_mode:
        mode_label = "REPORT SELECTED (input files only)"
    elif report_mode:
        mode_label = "REPORT ONLY"
    elif purge_mode:
        mode_label = "PURGE DATE RANGE then re-seed"
    elif prune_closed_mode:
        mode_label = "PRUNE CLOSED MERCHANTS"
    elif batch_seed_mode:
        mode_label = "BATCH SEED (list_new + list_update xlsx)"
    elif add_merchants_mode:
        mode_label = "ADD NEW MERCHANTS (no-touch existing)"
    elif append_mode:
        mode_label = "APPEND (continue from last transaction date)"
    elif reset_mode:
        mode_label = "FULL RESET"
    else:
        mode_label = "INITIAL SEED"

    print("=" * 60)
    print("  EDC POC Seeder")
    print(f"  Mode       : {mode_label}")
    print(f"  Date range : {DATE_START} — {DATE_END}")
    print(f"  CSV        : {CSV_PATH}")
    print(f"  Database   : {DB_CONFIG['dbname']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}")
    print("=" * 60)

    _ensure_merchant_columns(conn)

    if report_mode or report_selected_mode or append_mode or batch_seed_mode:
        _sync_merchant_data(conn)

    if prune_closed_mode:
        print("[1/3] Loading CSV for closure dates...")
        with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
            csv_rows = [r for r in csv.DictReader(f) if r.get("status", "").strip() == "ACTIVE"]
        print(f"      {len(csv_rows)} active POIs loaded.")

        print("[2/3] Pruning closed merchant data...")
        prune_closed_merchants(conn, csv_rows)

        print("[3/3] Loading merchants from DB for report...")
        merchants_info = load_merchants_from_db(conn, csv_rows)
        print(f"      {len(merchants_info)} merchants loaded.")

        print("\nGenerating merchant activity report...")
        generate_merchant_status_report(conn, merchants_info)
        conn.close()
        return

    if append_mode:
        print("[1/4] Loading CSV + batch xlsx for schedule data...")
        with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
            csv_rows = [r for r in csv.DictReader(f) if r.get("status", "").strip() == "ACTIVE"]
        _n_xlsx = _load_xlsx_supplement(csv_rows)
        print(f"      {len(csv_rows) - _n_xlsx} CSV POIs + {_n_xlsx} xlsx rows loaded.")

        append_start = get_append_start_date(conn)
        if append_start is None:
            print("      No existing transactions found — run a full seed first.")
            conn.close()
            return
        if append_start > DATE_END:
            print(f"      Already up to date (last date: {append_start - timedelta(days=1)}).")
            conn.close()
            return
        print(f"      Appending from {append_start} → {DATE_END} "
              f"({(DATE_END - append_start).days + 1} days).")

        print("[2/4] Loading reference data (QRIS issuers)...")
        qris_issuer_ids = load_qris_issuers(conn)

        print("[3/4] Loading merchants + cards from DB...")
        merchants_info = load_merchants_from_db(conn, csv_rows)
        card_ids       = load_cards_from_db(conn)
        print(f"      {len(merchants_info)} merchants, {len(card_ids)} cards.")

        print("[4/4] Generating and inserting transactions...")
        reset_trace_seq(conn)
        generate_and_insert_transactions(
            conn, merchants_info, card_ids, qris_issuer_ids,
            start_date=append_start, end_date=DATE_END,
        )

        print("\nRow count verification:")
        with conn.cursor() as cur:
            for table in ["transactions", "settlement", "transaction_log"]:
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                n = cur.fetchone()[0]
                print(f"      {table:<22} {n:>10,}")

        print("\nGenerating merchant activity report...")
        generate_merchant_status_report(conn, merchants_info)
        conn.close()
        return

    if report_mode:
        # Skip all seeding — load existing merchants from DB and generate report only
        print("[1/2] Loading CSV + batch xlsx for schedule data...")
        with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
            csv_rows = [r for r in csv.DictReader(f) if r.get("status", "").strip() == "ACTIVE"]
        _n_xlsx = _load_xlsx_supplement(csv_rows)
        print(f"      {len(csv_rows) - _n_xlsx} CSV POIs + {_n_xlsx} xlsx rows loaded.")

        print("[2/2] Loading merchants from DB...")
        merchants_info = load_merchants_from_db(conn, csv_rows)
        print(f"      {len(merchants_info)} merchants loaded.")

        print("\nGenerating merchant activity report...")
        generate_merchant_status_report(conn, merchants_info)
        conn.close()
        return

    if report_selected_mode:
        # Same as --report but restricted to merchants present in the input files
        print("[1/2] Loading CSV + batch xlsx for schedule data...")
        with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
            csv_rows = [r for r in csv.DictReader(f) if r.get("status", "").strip() == "ACTIVE"]
        _n_xlsx = _load_xlsx_supplement(csv_rows)
        print(f"      {len(csv_rows) - _n_xlsx} CSV POIs + {_n_xlsx} xlsx rows loaded.")

        print("[2/2] Loading merchants from DB...")
        merchants_info = load_merchants_from_db(conn, csv_rows)
        input_names = {r.get("name1", "").strip() for r in csv_rows if r.get("name1", "").strip()}
        merchants_info = [m for m in merchants_info if m["merchant_name"] in input_names]
        print(f"      {len(merchants_info)} merchants matched to input files.")

        print("\nGenerating merchant activity report (selected)...")
        generate_merchant_status_report(conn, merchants_info)
        conn.close()
        return

    if batch_seed_mode:
        print("[1/7] Reading list_edc.csv...")
        batch_rows: list[dict] = []
        rt_rows:    list[dict] = []
        if _BATCH_UNIFIED_CSV.exists():
            rows_u     = _read_unified_csv(_BATCH_UNIFIED_CSV)
            rt_rows    = [r for r in rows_u if r.get("batch_type") == "xlsx-realtime"]
            batch_rows = [r for r in rows_u if r.get("batch_type") != "xlsx-realtime"]
            del_rows   = [r for r in batch_rows if r.get("batch_type") == "xlsx-delete"]
            print(f"      list_edc.csv : {len(batch_rows)} batch rows, {len(rt_rows)} realtime")
        else:
            print(f"      list_edc.csv : not found — nothing to do.")
            conn.close()
            return
        if rt_rows:
            import csv as _csv_mod
            with open(_REALTIME_CSV, "w", newline="", encoding="utf-8") as _rtf:
                _w = _csv_mod.DictWriter(_rtf, fieldnames=["id", "poi_nm", "last_signal", "category", "street"])
                _w.writeheader()
                for _r in rt_rows:
                    _w.writerow({
                        "id":          _r.get("_xlsx_id", ""),
                        "poi_nm":      _r.get("name1", ""),
                        "last_signal": _r.get("last_signal", ""),
                        "category":    _r.get("primarycategorynm", ""),
                        "street":      _r.get("streetname", ""),
                    })
            print(f"      list_realtime.csv   : {len(rt_rows)} rows written")

        if not batch_rows:
            print("      No batch rows loaded — nothing to do.")
            conn.close()
            return

        print("[2/7] Identifying merchants not yet in DB...")
        new_to_db = filter_new_csv_rows(conn, batch_rows)
        print(f"      {len(new_to_db)} new (not in DB), "
              f"{len(batch_rows) - len(new_to_db)} already exist.")

        print("[3/7] Loading reference data...")
        acquirer_ids    = load_acquirers(conn)
        qris_issuer_ids = load_qris_issuers(conn)

        if new_to_db:
            idx_offset = get_merchant_idx_offset(conn)
            print(f"      Sequence offset: {idx_offset}")

            print("[4/7] Inserting admin areas (idempotent)...")
            area_map = insert_admin_areas(conn, new_to_db)

            print("[5/7] Inserting new merchants + terminals...")
            new_merchants_info = insert_merchants_and_terminals(
                conn, new_to_db, acquirer_ids, area_map, idx_offset=idx_offset
            )
            print(f"      {len(new_merchants_info)} merchants and terminals inserted.")
        else:
            new_merchants_info = []
            print("[4/7] No new merchants — skipping insert.")
            print("[5/7] (skipped)")

        # Also seed transactions for batch merchants already in DB but with zero transactions
        with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
            csv_base = [r for r in csv.DictReader(f) if r.get("status", "").strip() == "ACTIVE"]
        all_schedule_rows = csv_base + batch_rows
        all_batch_info = load_merchants_from_db(conn, all_schedule_rows)

        batch_names = {r.get("name1", "").strip() for r in batch_rows}
        batch_info_only = [m for m in all_batch_info if m["merchant_name"] in batch_names]

        # Prune post-closure transactions for delete merchants
        del_names = {r.get("name1", "").strip() for r in del_rows}
        del_info  = [m for m in all_batch_info if m["merchant_name"] in del_names]
        if del_info:
            print(f"      Pruning post-closure transactions for {len(del_info)} deleted merchant(s)...")
            with conn.cursor() as cur:
                for m_info in del_info:
                    cf = m_info.get("closed_from")
                    if not cf:
                        continue
                    close_ts = datetime.combine(cf, time(0, 0))
                    cur.execute(
                        "DELETE FROM transaction_log WHERE transaction_id IN "
                        "(SELECT transaction_id FROM transactions "
                        " WHERE merchant_id = %s AND transaction_time >= %s)",
                        (m_info["merchant_id"], close_ts),
                    )
                    cur.execute(
                        "DELETE FROM transactions WHERE merchant_id = %s AND transaction_time >= %s",
                        (m_info["merchant_id"], close_ts),
                    )
                    cur.execute(
                        "DELETE FROM settlement WHERE merchant_id = %s AND settlement_date >= %s",
                        (m_info["merchant_id"], cf),
                    )
            conn.commit()

        # Find batch merchants that have no transactions yet (checked after prune)
        with conn.cursor() as cur:
            cur.execute("SELECT merchant_id FROM transactions GROUP BY merchant_id")
            has_txn = {r[0] for r in cur.fetchall()}
        no_txn_info = [m for m in batch_info_only if m["merchant_id"] not in has_txn]

        if new_merchants_info or no_txn_info:
            to_seed = new_merchants_info + [m for m in no_txn_info
                                            if m["merchant_id"] not in
                                            {x["merchant_id"] for x in new_merchants_info}]
            print(f"[6/7] Generating transactions for {len(to_seed)} batch merchant(s) with no data...")
            card_ids = load_cards_from_db(conn)
            reset_trace_seq(conn)
            generate_and_insert_transactions(
                conn, to_seed, card_ids, qris_issuer_ids,
                start_date=DATE_START, end_date=DATE_END,
            )
        else:
            print("[6/7] All batch merchants already have transactions — skipping.")

        print("\nRow count verification:")
        with conn.cursor() as cur:
            for table in ["merchants", "terminals", "transactions", "settlement", "transaction_log"]:
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                n = cur.fetchone()[0]
                print(f"      {table:<22} {n:>10,}")

        print("[7/7] Generating report (all merchants, batch-tagged)...")
        # Re-load merchants_info so new transaction data is reflected in the report
        all_merchants_info = load_merchants_from_db(conn, all_schedule_rows)
        generate_merchant_status_report(conn, all_merchants_info)
        conn.close()
        return

    if add_merchants_mode:
        print("[1/6] Loading CSV...")
        with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
            csv_rows = [r for r in csv.DictReader(f)
                        if r.get("status", "").strip() == "ACTIVE"]
        print(f"      {len(csv_rows)} active POIs in CSV.")

        print("[2/6] Identifying new merchants...")
        new_rows = filter_new_csv_rows(conn, csv_rows)
        if not new_rows:
            print("      No new merchants found — DB is already up to date.")
            conn.close()
            return
        print(f"      {len(new_rows)} new merchant(s): "
              + ", ".join(r.get("name1", "") for r in new_rows))

        idx_offset = get_merchant_idx_offset(conn)
        print(f"      Sequence offset: {idx_offset}")

        print("[3/6] Inserting admin areas (idempotent)...")
        acquirer_ids    = load_acquirers(conn)
        qris_issuer_ids = load_qris_issuers(conn)
        area_map = insert_admin_areas(conn, new_rows)

        print("[4/6] Inserting new merchants + terminals...")
        new_merchants_info = insert_merchants_and_terminals(
            conn, new_rows, acquirer_ids, area_map, idx_offset=idx_offset
        )
        print(f"      {len(new_merchants_info)} merchants and terminals inserted.")

        print("[5/6] Loading cards + generating transactions for new merchants...")
        card_ids = load_cards_from_db(conn)
        reset_trace_seq(conn)
        generate_and_insert_transactions(
            conn, new_merchants_info, card_ids, qris_issuer_ids,
            start_date=DATE_START, end_date=DATE_END,
        )

        print("\nRow count verification:")
        with conn.cursor() as cur:
            for table in ["merchants", "terminals", "transactions", "settlement", "transaction_log"]:
                cur.execute(f"SELECT COUNT(*) FROM {table}")
                n = cur.fetchone()[0]
                print(f"      {table:<22} {n:>10,}")

        print("[6/6] Generating report (all merchants)...")
        all_merchants_info = load_merchants_from_db(conn, csv_rows)
        generate_merchant_status_report(conn, all_merchants_info)
        conn.close()
        return

    print("[2/7] Loading reference data (acquirers + QRIS issuers)...")
    acquirer_ids    = load_acquirers(conn)
    qris_issuer_ids = load_qris_issuers(conn)
    print(f"      {len(acquirer_ids)} acquirers, {len(qris_issuer_ids)} QRIS issuers.")

    print("[3/7] Loading CSV...")
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        csv_rows = [r for r in csv.DictReader(f) if r.get("status", "").strip() == "ACTIVE"]
    print(f"      {len(csv_rows)} active POIs loaded.")

    if purge_mode:
        print("[1/7] Purging transaction data for date range...")
        purge_date_range(conn)

        print("[4/7] Loading admin areas from DB (skipped — purge mode)...")
        print("[5/7] Loading merchants + terminals from DB...")
        merchants_info = load_merchants_from_db(conn, csv_rows)
        print(f"      {len(merchants_info)} merchants loaded from DB.")

        print("[6/7] Loading cards from DB (skipped — purge mode)...")
        card_ids = load_cards_from_db(conn)
        print(f"      {len(card_ids)} cards loaded from DB.")
    else:
        print("[1/7] Resetting database...")
        reset_database(conn)

        print("[4/7] Inserting admin areas (Province → City → District)...")
        area_map = insert_admin_areas(conn, csv_rows)
        with conn.cursor() as cur:
            cur.execute("SELECT area_level, COUNT(*) FROM admin_areas GROUP BY 1 ORDER BY 1")
            for lvl, cnt in cur.fetchall():
                label = {1: "Province", 2: "City/Regency", 3: "District", 4: "Village"}[lvl]
                print(f"      Level {lvl} ({label}): {cnt}")

        print("[5/7] Inserting merchants + terminals...")
        merchants_info = insert_merchants_and_terminals(conn, csv_rows, acquirer_ids, area_map)
        print(f"      {len(merchants_info)} merchants and terminals inserted.")

        print("[6/7] Inserting cards...")
        card_ids = insert_cards(conn, count=300)
        print(f"      {len(card_ids)} cards inserted.")

    print("[7/7] Generating transactions, settlement, and audit logs...")
    generate_and_insert_transactions(conn, merchants_info, card_ids, qris_issuer_ids)

    print("\nRow count verification:")
    with conn.cursor() as cur:
        for table in [
            "admin_areas", "merchants", "terminals", "cards", "qris_issuers",
            "transactions", "settlement", "transaction_log",
        ]:
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            n = cur.fetchone()[0]
            print(f"      {table:<22} {n:>10,}")

    print("\n[8/8] Merchant activity report...")
    generate_merchant_status_report(conn, merchants_info)

    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
